import os
import re
from io import BytesIO
from flask import Flask, request, render_template, send_file, jsonify, flash, redirect, url_for
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

app = Flask(__name__, template_folder='templates', static_folder='static')
app.secret_key = os.environ.get('SECRET_KEY') or 'dev-secret-key'

# Configure upload folder
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # 16MB limit

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def clean_filename(name, options=None):
    """Clean filename with customizable character handling"""
    if not name or pd.isna(name):
        return {'cleaned': '', 'removed': []}
    
    options = options or {}
    original = str(name).strip()
    removed_chars = []
    
    # Space handling
    space_option = options.get('space_replacement', 'remove')
    if space_option == 'underscore':
        original = original.replace(' ', '_')
    elif space_option == 'hyphen':
        original = original.replace(' ', '-')
    
    # Case handling
    case_option = options.get('text_case', 'original')
    if case_option == 'lower':
        original = original.lower()
    elif case_option == 'upper':
        original = original.upper()
    elif case_option == 'title':
        original = original.title()
    
    # Character removal
    chars_to_remove = []
    special_chars = [
    ('/', 'keep_slash'),
    ('\\', 'keep_backslash'),
    (':', 'keep_colon'),
    ('*', 'keep_asterisk'),
    ('?', 'keep_question'),
    ('"', 'keep_dquote'),
    ('<', 'keep_ltgt'),
    ('>', 'keep_ltgt'),
    ('|', 'keep_pipe')
]

    
    for char, option in special_chars:
        if not options.get(option, True):
            chars_to_remove.append(char)
    
    # Always remove control characters
    chars_to_remove.extend(['\\x00-\\x1F'])
    
    # Remove duplicate symbols if enabled
    if options.get('remove_repeats'):
        original = re.sub(r'([-_\.])\1+', r'\1', original)
    
    pattern = f'[{"".join(chars_to_remove)}]' if chars_to_remove else None
    
    cleaned = original
    if pattern:
        removed_chars = list(set(re.findall(pattern, cleaned)))
        cleaned = re.sub(pattern, '', cleaned).strip()
    
    return {
        'original': str(name).strip(),
        'cleaned': cleaned,
        'removed': removed_chars
    }

@app.route('/', methods=['GET'])
def home():
    return render_template('index.html')

@app.route('/preview', methods=['POST'])
def preview():
    data = request.json
    result = clean_filename(data.get('text', ''), data.get('options', {}))
    return jsonify(result)

@app.route('/process', methods=['POST'])
def process_files():
    if 'files[]' not in request.files:
        flash('No files uploaded')
        return redirect(url_for('home'))
    
    files = request.files.getlist('files[]')
    options = request.form.to_dict()
    results = []
    
    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            try:
                if filename.endswith(('.xlsx', '.xls')):
                    from openpyxl.utils import column_index_from_string
                    wb = openpyxl.load_workbook(filepath, read_only=False)
                    sheet_name = options.get('sheet')
                    sheet = wb[sheet_name] if sheet_name else wb.active
                    col_letter = options.get('column', 'A')
                    col_idx = column_index_from_string(col_letter)
                    start_row = int(options.get('start_row', 1))
                    end_row = int(options.get('end_row', sheet.max_row))
                    
                    for row in sheet.iter_rows(
                        min_row=start_row,
                        max_row=end_row,
                        min_col=col_idx,
                        max_col=col_idx,
                        values_only=True
                    ):
                        if row[0] and str(row[0]).strip():
                            result = clean_filename(row[0], options)
                            results.append(result)
                
                elif filename.endswith('.csv'):
                    df = pd.read_csv(filepath)
                    if not df.empty:
                        for value in df.iloc[:, 0].astype(str):
                            if value.strip():
                                result = clean_filename(value, options)
                                results.append(result)
            
            except Exception as e:
                flash(f"Error processing {filename}: {str(e)}")
            finally:
                os.remove(filepath)
    
    if not results:
        flash("No valid filenames found to process")
        return redirect(url_for('home'))
    
    return render_template('results.html', results=results)

@app.route('/export', methods=['POST'])
def export():
    data = request.json
    format = data.get('format', 'excel')
    results = data.get('results', [])
    
    if not results:
        return jsonify({'error': 'No data to export'}), 400
    
    df = pd.DataFrame(results)
    
    if format == 'csv':
        output = df.to_csv(index=False)
        mimetype = 'text/csv'
        ext = 'csv'
    elif format == 'json':
        output = df.to_json(orient='records')
        mimetype = 'application/json'
        ext = 'json'
    else:
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ext = 'xlsx'
    
    return send_file(
        output if format == 'excel' else BytesIO(output.encode()),
        mimetype=mimetype,
        as_attachment=True,
        download_name=f'cleaned_filenames.{ext}'
    )

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
